from openpyxl import load_workbook
import numpy as np
import json
import os
from google import genai
from dotenv import load_dotenv
from pathlib import Path
from google.genai import types
import re


# =============================
# GESTION DES PERMISSIONS
# =============================

def detect_access_level_from_path(file_path: str) -> tuple:
    """
    Détecte le niveau d'accès associé à un chemin de fichier, en se basant sur des règles prédéfinies.

    Ce module permet d'analyser un chemin donné pour en extraire un certain niveau d'accès,
    un type de document et une liste de permissions requises. Il traite à la fois les dossiers
    clés et cherche des correspondances avec des mots-clés dans les chemins normalisés.

    :param file_path: Le chemin complet du fichier à analyser.
    :type file_path: str
    :return: Une tupla contenant le niveau d'accès détecté (e.g. "public", "internal", "confidential"),
             le type de document associé (e.g. "public_data", "admin_data") ainsi que la liste
             des permissions nécessaires pour accéder au document.
    :rtype: tuple
    :raises ValueError: Une exception si le chemin fourni est invalide ou mal formé.
    """

    # Normaliser le chemin
    normalized_path = Path(file_path).as_posix().lower()
    path_parts = Path(file_path).parts

    # Mapping des dossiers vers les niveaux d'accès
    folder_mappings = {
        # Niveaux d'accès principaux
        'public': {
            'access_level': 'public',
            'document_type': 'public_data',
            'required_permissions': ['read_public_docs']
        },
        'interne': {
            'access_level': 'internal',
            'document_type': 'internal_data',
            'required_permissions': ['read_internal_docs', 'read_public_docs']
        },
        'internal': {
            'access_level': 'internal',
            'document_type': 'internal_data',
            'required_permissions': ['read_internal_docs', 'read_public_docs']
        },
        'admin': {
            'access_level': 'confidential',
            'document_type': 'admin_data',
            'required_permissions': ['read_confidential_docs', 'read_internal_docs', 'read_public_docs']
        },
        'confidential': {
            'access_level': 'confidential',
            'document_type': 'confidential_data',
            'required_permissions': ['read_confidential_docs', 'read_internal_docs', 'read_public_docs']
        },
        'confidentiel': {
            'access_level': 'confidential',
            'document_type': 'confidential_data',
            'required_permissions': ['read_confidential_docs', 'read_internal_docs', 'read_public_docs']
        },
        # Types de documents spécifiques
        'statistiques': {
            'access_level': 'internal',
            'document_type': 'statistics',
            'required_permissions': ['read_internal_docs', 'read_public_docs']
        },
        'finances': {
            'access_level': 'confidential',
            'document_type': 'financial',
            'required_permissions': ['read_confidential_docs', 'read_internal_docs', 'read_public_docs']
        },
        'rh': {
            'access_level': 'internal',
            'document_type': 'hr_data',
            'required_permissions': ['read_internal_docs', 'read_public_docs']
        },
        'strategie': {
            'access_level': 'confidential',
            'document_type': 'strategy',
            'required_permissions': ['read_confidential_docs', 'read_internal_docs', 'read_public_docs']
        }
    }

    # Chercher dans les parties du chemin
    for part in path_parts:
        part_lower = part.lower()
        if part_lower in folder_mappings:
            mapping = folder_mappings[part_lower]
            print(f"Détecté niveau d'accès '{mapping['access_level']}' pour dossier '{part}' dans {file_path}")
            return mapping['access_level'], mapping['document_type'], mapping['required_permissions']

    # Fallback : chercher dans le chemin complet pour des mots-clés
    for keyword, mapping in folder_mappings.items():
        if keyword in normalized_path:
            print(f"Détecté niveau d'accès '{mapping['access_level']}' par mot-clé '{keyword}' dans {file_path}")
            return mapping['access_level'], mapping['document_type'], mapping['required_permissions']

    # Par défaut : public
    print(f"Aucun niveau d'accès détecté pour {file_path}, utilisation de 'public' par défaut")
    return 'public', 'general', ['read_public_docs']


def enrich_metadata_with_permissions(metadata: dict, file_path: str) -> dict:
    """
    Cette fonction enrichit un dictionnaire de métadonnées prédéfini en y ajoutant des informations
    liées aux permissions ainsi que des données contextuelles extraites du chemin du fichier.

    :param metadata: Le dictionnaire contenant les métadonnées de base.
    :type metadata: dict
    :param file_path: Le chemin absolu du fichier utilisé pour déterminer les permissions
                     associées et les métadonnées complémentaires.
    :type file_path: str
    :return: Un dictionnaire contenant les métadonnées originales augmentées avec des champs
             relatifs aux permissions et des informations contextuelles.
    :rtype: dict
    """

    access_level, document_type, required_permissions = detect_access_level_from_path(file_path)

    # Enrichir les métadonnées
    enriched_metadata = {
        **metadata,
        # informations de permissions
        'access_level': access_level,
        'document_type': document_type,
        'required_permissions': ','.join(required_permissions),  # String pour ChromaDB
        'permissions_list': required_permissions,  # Liste pour usage interne
        # Informations de traçabilité
        'source_directory': str(Path(file_path).parent),
        'detected_from_path': True,
        'extraction_timestamp': str(Path().cwd())  # Pour debug
    }

    return enriched_metadata


# =============================
# 1. EXTRACTION DES TABLEAUX
# =============================

def sheet_to_presence_matrix(sheet):
    """
    Convertit une feuille de calcul en une matrice de présence.

    La fonction parcourt chaque cellule d'une feuille de calcul et crée une matrice
    de présence où chaque cellule contenant une valeur non nulle ou non vide est
    représentée par un `1` et les autres par un `0`.

    :param sheet: La feuille de calcul à convertir.
    :type sheet: openpyxl.worksheet.worksheet.Worksheet
    :return: Une matrice de présence (tableau Numpy) représentant les données
        de la feuille.
    :rtype: numpy.ndarray
    """
    max_row = sheet.max_row
    max_col = sheet.max_column
    matrix = np.zeros((max_row, max_col), dtype=int)

    for row in range(1, max_row + 1):
        for col in range(1, max_col + 1):
            value = sheet.cell(row=row, column=col).value
            if value is not None and str(value).strip() != "":
                matrix[row - 1, col - 1] = 1
    return matrix


from scipy.ndimage import label


def find_blocks(matrix):
    """
    Analyse une matrice pour identifier des blocs connectés et renvoie des informations
    sur les positions et les limites de ces blocs. Un bloc est défini comme un groupe
    de cellules connectées adjacentes (horizontalement, verticalement ou en diagonale).

    :param matrix: Une matrice binaire où les cellules valorisées à 1 représentent des
        éléments d'un bloc.
    :type matrix: numpy.ndarray
    :return: Une liste de dictionnaires représentant chaque bloc. Chaque dictionnaire
        contient les informations suivantes :
        - "min_row" : L'indice de la première ligne occupée par le bloc.
        - "max_row" : L'indice de la dernière ligne occupée par le bloc.
        - "min_col" : L'indice de la première colonne occupée par le bloc.
        - "max_col" : L'indice de la dernière colonne occupée par le bloc.
        - "positions" : Un tableau numpy contenant les positions (lignes, colonnes)
            de chaque cellule du bloc.
    :rtype: list[dict]
    """
    structure = np.ones((3, 3))
    labeled, num_features = label(matrix, structure=structure)
    blocks = []
    for block_id in range(1, num_features + 1):
        positions = np.argwhere(labeled == block_id)
        rows = positions[:, 0]
        cols = positions[:, 1]
        min_row, max_row = rows.min(), rows.max()
        min_col, max_col = cols.min(), cols.max()
        blocks.append({
            "min_row": min_row,
            "max_row": max_row,
            "min_col": min_col,
            "max_col": max_col,
            "positions": positions
        })
    return blocks


def extract_table(sheet, block):
    """
    Extrait une table d'une feuille donnée en fonction des limites définies par un bloc.
    La fonction lit les cellules correspondantes dans la feuille et retourne une liste
    représentant la table.

    :param sheet: Objet représentant la feuille depuis laquelle la table sera extraite.
                  Devrait fournir un accès aux cellules grâce à une méthode cell().
    :param block: Dictionnaire contenant les limites de la table à extraire. Les clefs
                  attendues sont "min_row", "max_row", "min_col", "max_col".
    :return: Liste représentant la table extraite. Chaque sous-liste représente une ligne
             de la table, remplie par les valeurs des cellules correspondantes.
    """
    table = []
    for r in range(block["min_row"] + 1, block["max_row"] + 2):
        row = []
        for c in range(block["min_col"] + 1, block["max_col"] + 2):
            row.append(sheet.cell(row=r, column=c).value)
        table.append(row)
    return table


def is_title_block(table):
    """
    Détermine si un tableau donné peut être considéré comme un bloc de titre.

    Cette fonction permet d'identifier si un tableau spécifique contient une rangée unique qui
    peut être interprétée comme un bloc de titre, en se basant sur les critères suivants :
    - Le tableau doit comporter une seule ligne.
    - Chaque cellule de cette ligne, si elle contient une chaîne, doit avoir une longueur
      minimale une fois les espaces inutiles supprimés.

    :param table: Liste bidimensionnelle représentant un tableau. Chaque entrée de la liste
                  externe représente une ligne, et chaque élément dans cette ligne représente
                  une cellule.
    :type table: list[list[Any]]
    :return: Retourne True si le tableau est un bloc de titre, sinon False.
    :rtype: bool
    """
    if len(table) == 1:
        if any(isinstance(cell, str) and len(cell.strip()) > 10 for cell in table[0]):
            return True
    return False


def assign_titles_to_tables(sheet, blocks):
    """
    Assigne des titres aux tableaux extraits d'une feuille à partir des blocs
    fournis en paramètre. Cette fonction analyse et traite les blocs pour associer
    un contexte de titre à chaque tableau correspondant.

    :param sheet: La feuille de calcul (workbook) à analyser, contenant
        les données sources.
    :type sheet: Any
    :param blocks: Une liste de blocs décrivant les zones d’intérêt dans la
        feuille. Chaque bloc doit inclure des clés telles que 'min_row',
        'max_row', 'min_col' et 'max_col'.
    :type blocks: list[dict]
    :return: Une liste de dictionnaires représentant chaque tableau avec :
        - 'titre': Le titre contextuel associé au tableau, s'il est disponible.
        - 'tableau': Le tableau extrait de la feuille de calcul.
        - 'coordonnees': Les coordonnées (row et col) des limites des zones du
          tableau.
    :rtype: list[dict]
    """
    title_context = None
    result_tables = []
    for idx, block in enumerate(blocks):
        table = extract_table(sheet, block)
        if is_title_block(table):
            title_context = " ".join([str(cell) for cell in table[0] if cell])
            continue  # on saute le stockage des titres seuls
        else:
            result_tables.append({
                "titre": title_context,
                "tableau": table,
                "coordonnees": (block['min_row'] + 1, block['max_row'] + 1, block['min_col'] + 1, block['max_col'] + 1)
            })
    return result_tables


def safe_filename(s):
    """
    Renvoie une version sécurisée d'une chaîne donnée à utiliser comme nom de fichier.
    Supprime ou remplace les caractères non valides afin de garantir la compatibilité du nom avec
    les systèmes de fichiers.

    :param s: La chaîne d'entrée à sécuriser.
    :type s: str
    :return: Une version sécurisée de la chaîne d'entrée contenant uniquement des caractères alphanumériques,
             des espaces, des points (.), des traits d'union (-) et des traits de soulignement (_).
    :rtype: str
    """
    return "".join(c if c.isalnum() or c in " ._-" else "_" for c in str(s))


def save_table_json(table, meta, output_dir, idx):
    """
    Sauvegarde une table et ses métadonnées au format JSON dans un répertoire
    donné. Cette fonction génère un fichier JSON spécifique pour chaque table
    et inclut un indicateur visuel basé sur le niveau d'accès des métadonnées.

    :param table: Les données de la table à sauvegarder dans le fichier JSON.
    :type table: dict
    :param meta: Les métadonnées associées à la table, contenant au minimum le
                 champ "access_level" avec des valeurs possibles telles que
                 "public", "internal" ou "confidential".
    :type meta: dict
    :param output_dir: Le répertoire où le fichier JSON doit être sauvegardé.
    :type output_dir: str
    :param idx: Indice utilisé pour nommer le fichier, sous un format à trois
                chiffres (ex : 001, 002...).
    :type idx: int
    :return: Chemin complet du fichier JSON généré.
    :rtype: str
    """
    filename = f"tableau_{idx:03d}.json"
    path = os.path.join(output_dir, filename)

    # Ajouter un indicateur visuel du niveau d'accès dans le JSON
    access_indicator = {
        'public': '[PUBLIC]',
        'internal': '[INTERNE]',
        'confidential': '[CONFIDENTIEL]'
    }.get(meta.get('access_level', 'public'), '[DOCUMENT]')

    json_data = {
        "tableau": table,
        "access_indicator": access_indicator,  # Pour faciliter la lecture
        **meta
    }

    with open(path, "w", encoding="utf-8") as f:
        json.dump(json_data, f, ensure_ascii=False, indent=2)

    print(f"Sauvegardé: {access_indicator} {filename}")
    return path


def full_extraction_to_json(xlsx_path, output_base):
    """
    Effectue l'extraction complète de données d'un fichier Excel au format JSON dans un
    répertoire de base donné. Les feuilles, tableaux, et métadonnées associées sont
    analysés et sauvegardés en conséquence. En outre, un index global est généré, avec
    des statistiques sur le niveau d'accès détecté.

    :param xlsx_path: Chemin vers le fichier Excel source
    :param output_base: Répertoire de base où les fichiers JSON extraits seront sauvegardés
    :return: Aucun
    """
    wb = load_workbook(xlsx_path, data_only=True)
    base_name = os.path.splitext(os.path.basename(xlsx_path))[0]
    output_file_dir = os.path.join(output_base, safe_filename(base_name))
    os.makedirs(output_file_dir, exist_ok=True)
    global_index = []

    # Détecter le niveau d'accès du fichier
    access_level, document_type, required_permissions = detect_access_level_from_path(xlsx_path)
    print(f"\nFichier {xlsx_path}:")
    print(f"  Niveau d'accès détecté: {access_level}")
    print(f"  Type de document: {document_type}")
    print(f"  Permissions requises: {required_permissions}")

    for sheetname in wb.sheetnames:
        sheet = wb[sheetname]
        matrix = sheet_to_presence_matrix(sheet)
        blocks = find_blocks(matrix)
        tables = assign_titles_to_tables(sheet, blocks)
        feuille_dir = os.path.join(output_file_dir, safe_filename(sheetname))
        os.makedirs(feuille_dir, exist_ok=True)

        for idx, t in enumerate(tables, 1):
            # Métadonnées de base
            base_meta = {
                "fichier_source": base_name,
                "nom_feuille": sheetname,
                "range_bloc": f"{t['coordonnees']}",
                "titre_contextuel": t['titre'],
            }

            enriched_meta = enrich_metadata_with_permissions(base_meta, xlsx_path)

            table_json_path = save_table_json(
                t['tableau'], enriched_meta, feuille_dir, idx
            )

            # Index global avec informations de permissions
            global_index.append({
                "fichier_source": base_name,
                "nom_feuille": sheetname,
                "titre_contextuel": t['titre'],
                "tableau_json": os.path.relpath(table_json_path, output_base),
                "access_level": access_level,
                "document_type": document_type,
                "required_permissions": ','.join(required_permissions),
                "permissions_summary": f"{access_level.upper()} - {document_type}"
            })

    # Sauvegarde de l'index global pour la recherche
    index_path = os.path.join(output_base, "index.json")
    with open(index_path, "w", encoding="utf-8") as idxf:
        json.dump(global_index, idxf, ensure_ascii=False, indent=2)

    # Statistiques par niveau d'accès
    stats_by_level = {}
    for item in global_index:
        level = item['access_level']
        stats_by_level[level] = stats_by_level.get(level, 0) + 1

    print(f"\nExtraction terminée ! {len(global_index)} tableaux enregistrés.")
    print("Répartition par niveau d'accès:")
    for level, count in stats_by_level.items():
        indicator = {'public': '[PUBLIC]', 'internal': '[INTERNE]', 'confidential': '[CONFIDENTIEL]'}.get(level,
                                                                                                          '[AUTRE]')
        print(f"  {indicator} {level}: {count} tableaux")

    print(f"Index sauvegardé dans: {index_path}")


# =============================
# 2. EXTRACTION DES PDFs
# =============================
def extract_pdf_text(pdf_path):
    """
    Extrait le texte des fichiers PDF dans un répertoire donné, en résumant leur contenu et
    enrichissant les données avec des métadonnées comme le niveau d'accès, le type de document et
    les permissions nécessaires.

    Cette fonction analyse chaque fichier PDF présent dans le chemin fourni, génère un résumé
    structuré en format JSON, et l'enregistre sur disque avec les métadonnées pertinentes.

    :param pdf_path: Chemin du répertoire à analyser. Seuls les fichiers PDF présents dans ce
        répertoire seront traités.
    :type pdf_path: str
    :return: Une chaîne indiquant que l'extraction des PDF est terminée.
    :rtype: str
    :raises Exception: Si une erreur survient lors du traitement des fichiers PDF ou de la
        génération des résumés.
    """
    client = genai.Client(api_key="AIzaSyAlgDF_78gtHpOVYcQ5-ucU6wp47vIc8Ns")
    data_file = Path(pdf_path)

    for file in data_file.iterdir():
        if file.suffix == ".pdf":
            access_level, document_type, required_permissions = detect_access_level_from_path(str(file))

            prompt = (f"Voici le chemin vers le fichier PDF : {file}"
                      "Fait un résumé de ce fichier PDF"
                      "Produit un JSON respectant la forme suivante :"
                      "Resume = { 'fichier_source': string, 'resume': string, 'chemin': string, "
                      f"'access_level': '{access_level}', 'document_type': '{document_type}', "
                      f"'required_permissions': '{','.join(required_permissions)}' }}"
                      "Return: Resume")

            try:
                response = client.models.generate_content(
                    model="gemini-2.5-flash",
                    contents=[
                        types.Part.from_bytes(
                            data=file.read_bytes(),
                            mime_type='application/pdf',
                        ),
                        prompt])

                # Enrichir la réponse avec les métadonnées de permissions
                texte = nettoyer_json(response.text)

                # Ajouter les informations de permissions si elles manquent
                try:
                    resume_data = json.loads(texte)
                    resume_data.update({
                        'access_level': access_level,
                        'document_type': document_type,
                        'required_permissions': ','.join(required_permissions),
                        'pdf_path': str(file),
                        'extraction_method': 'gemini_ai'
                    })
                    texte = json.dumps(resume_data, ensure_ascii=False, indent=2)
                except json.JSONDecodeError:
                    print(f"Erreur parsing JSON pour {file}, utilisation texte brut")

                output_file = f"output/index_{file.stem}_{access_level}.json"
                with open(output_file, "w", encoding="utf-8") as f:
                    f.write(texte)

                access_indicator = {'public': '[PUBLIC]', 'internal': '[INTERNE]',
                                    'confidential': '[CONFIDENTIEL]'}.get(access_level, '[PDF]')
                print(f"PDF traité: {access_indicator} {file.name} -> {output_file}")

            except Exception as e:
                print(f"Erreur traitement PDF {file}: {e}")

    return "Extraction PDF terminée"


def nettoyer_json(texte):
    """
    Nettoie un texte en supprimant les délimitations spécifiques des blocs
    de code JSON (```json et les ```) ainsi que les espaces ou lignes vides
    supplémentaires autour du contenu.

    Cela permet d'extraire uniquement le contenu JSON brut pour un traitement
    ultérieur.

    :param texte: Chaîne de caractères contenant potentiellement un bloc de
        code JSON entouré de délimitations ```json ou ``` à nettoyer.
    :returns: Chaîne de caractères représentant le contenu JSON nettoyé,
        sans délimitations ni espaces superflus.
    """
    pattern = r'^```json\s*\n?|```\s*$'
    texte = re.sub(pattern, '', texte, flags=re.MULTILINE).strip()
    return texte


def analyze_directory_structure(data_dir: str):
    """
    Analyse la structure d'un répertoire pour catégoriser les fichiers Excel (.xlsx) en
    fonction de leur niveau d'accès, type de document, et permissions requises.

    Cette fonction scanne le répertoire spécifié et détecte les fichiers Excel. Elle analyse
    chaque fichier pour déterminer son niveau d'accès, son type et ses attributs, puis regroupe
    ces informations dans un dictionnaire structuré. Elle affiche également une vue d'ensemble
    de cette structure avec le nombre de fichiers pour chaque niveau d'accès.

    :param data_dir: Chemin du répertoire à analyser
    :type data_dir: str
    :returns: Un dictionnaire contenant les fichiers catégorisés par niveau d'accès.
        Chaque clé représente un niveau d'accès et associe une liste de dictionnaires contenant
        des informations détaillées sur les fichiers.
    :rtype: dict[str, list[dict[str, str]]]
    """
    data_path = Path(data_dir)

    if not data_path.exists():
        print(f"Dossier {data_dir} non trouvé")
        return

    print(f"Analyse de la structure: {data_path}")
    print("=" * 50)

    file_count_by_level = {}

    for file_path in data_path.rglob("*.xlsx"):
        access_level, document_type, required_permissions = detect_access_level_from_path(str(file_path))

        if access_level not in file_count_by_level:
            file_count_by_level[access_level] = []

        file_count_by_level[access_level].append({
            'file': file_path.name,
            'path': str(file_path.relative_to(data_path)),
            'document_type': document_type
        })

    # Afficher les résultats
    for level, files in file_count_by_level.items():
        indicator = {'public': '[PUBLIC]', 'internal': '[INTERNE]', 'confidential': '[CONFIDENTIEL]'}.get(level,
                                                                                                          '[AUTRE]')
        print(f"\n{indicator} Niveau {level.upper()} ({len(files)} fichiers):")
        for file_info in files:
            print(f"  - {file_info['path']} ({file_info['document_type']})")

    return file_count_by_level


def extract_all_files_in_directory(data_dir: str, output_base: str = "output"):
    """
    Extrait tous les fichiers Excel d'un répertoire donné, les traite et enregistre les résultats
    au format JSON dans un répertoire de sortie spécifié. Cette fonction recherche
    de manière récursive les fichiers ayant une extension `.xlsx` dans le répertoire de données.

    L'extraction de chaque fichier est réalisée en appelant une autre fonction,
    et tout problème lors du traitement des fichiers individuels sera signalé
    sans interrompre l'exécution de la fonction pour les autres fichiers.

    :param data_dir: Chemin du répertoire contenant les fichiers Excel à extraire.
    :param output_base: Chemin de base où les fichiers JSON extraits seront générés. Par défaut "output".
    :return: Aucun retour de valeur. Les résultats sont enregistrés dans le répertoire de sortie.
    """
    data_path = Path(data_dir)

    if not data_path.exists():
        print(f"Dossier {data_dir} non trouvé")
        return

    print(f"Extraction de tous les fichiers dans: {data_path}")

    excel_files = list(data_path.rglob("*.xlsx"))
    print(f"Trouvé {len(excel_files)} fichiers Excel")

    for file_path in excel_files:
        print(f"\nTraitement: {file_path}")
        try:
            full_extraction_to_json(str(file_path), output_base)
        except Exception as e:
            print(f"Erreur avec {file_path}: {e}")

    print(f"\n Extraction complète terminée!")
    print(f"Résultats dans: {output_base}/")


if __name__ == "__main__":
    # Configuration
    os.chdir("/home/aissa/Bureau/Projet_Chatbot/chatbot_maroc/backend_python")

    # Exemple d'utilisation avec la nouvelle structure

    # 1. Analyser la structure des dossiers
    print("1. ANALYSE DE LA STRUCTURE:")
    analyze_directory_structure("data")

    # 2. Exemple d'extraction d'un fichier spécifique
    print("\n2. EXTRACTION D'UN FICHIER:")
    # xlsx_path = "data/public/stats_publiques.xlsx"  # Sera détecté comme public
    # xlsx_path = "data/internal/rapport_interne.xlsx"  # Sera détecté comme internal
    # xlsx_path = "data/admin/finances_confidentielles.xlsx"  # Sera détecté comme confidential

    xlsx_path = "data/admin/20230420_Capital Humain au Maroc_V-Final-2.xlsx"  # Sera détecté comme admin
    output_base = "output"
    full_extraction_to_json(xlsx_path, output_base)

    # 3. Extraction de tous les fichiers (optionnel)
    # print("\n3. EXTRACTION COMPLÈTE:")
    # extract_all_files_in_directory("data", "output")

    # 4. Extraction PDF avec permissions (optionnel)
    # print("\n4. EXTRACTION PDF:")
    # pdf_path = Path("data")
    # extract_pdf_text(pdf_path)